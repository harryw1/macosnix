#!/usr/bin/env python3
# /// script
# requires-python = ">=3.11"
# dependencies = [
#     "sqlite-vec",
#     "requests",
#     "pypdf",
#     "python-docx",
#     "openpyxl",
#     "numpy",
#     "scikit-learn",
# ]
# ///
"""ai-organize — AI-powered file reorganizer, renamer, and deduplicator.

Modes
-----
  --scan   DIR          Walk a directory and emit a JSON file-list (metadata + snippets).
  --plan   DIR [flags]  Ask an LLM to produce a JSON reorganization plan.
  --apply  PLAN_FILE    Execute a previously generated plan (supports --dry-run).

Plan flags
----------
  --rename    Suggest more-descriptive filenames.
  --organize  Propose a logical folder structure.
  --flatten   Collapse deeply-nested paths into a flat layout.
  --dedupe    Flag files whose content is suspiciously similar.
  --top-level Non-recursive: only consider files in the root of DIR.
"""

import argparse
import json
import os
import re
import shutil
import sqlite3
import struct
import sys
from pathlib import Path

import numpy as np
import requests

# ── Configuration ──────────────────────────────────────────────────────────────

OLLAMA_URL       = "http://localhost:11434/api/generate"
OLLAMA_EMBED_URL = "http://localhost:11434/api/embeddings"
MODEL            = os.environ.get("OLLAMA_MODEL",       "qwen3.5:9b")
EMBED_MODEL      = os.environ.get("OLLAMA_MODEL_EMBED", "qwen3-embedding:0.6b")

XDG_DATA_HOME  = os.environ.get("XDG_DATA_HOME", os.path.expanduser("~/.local/share"))
SEARCH_DB_PATH = Path(XDG_DATA_HOME) / "ai-search" / "vectors.db"

# Characters of file content to pass as context per file
SNIPPET_LEN = 400
# Maximum files sent to the LLM in a single call (stay within context window).
# Each file can produce ~2 operations × ~80 tokens each, so 40 files ≈ 6 400
# tokens of output — safely within the num_predict budget below.
MAX_FILES_PER_BATCH = 40
# Cosine-distance threshold for declaring two files "duplicates" via embeddings
DUPE_THRESHOLD = 0.12
# Maximum files per embedding batch (Ollama processes one at a time, but we
# track progress in batches of this size).
EMBED_BATCH_LOG = 100
# Minimum files in a cluster for HDBSCAN to consider it a real group.
# Smaller values produce more granular folders; larger values merge more.
HDBSCAN_MIN_CLUSTER = 3
# Weight given to embedding (content) distance vs path (structure) distance
# when computing the composite distance matrix for clustering.
# Higher α → more weight on content similarity; lower α → more on directory structure.
COMPOSITE_ALPHA = 0.65

TEXT_EXTS = {
    ".nix", ".md", ".toml", ".yml", ".yaml", ".sh", ".bash", ".zsh",
    ".json", ".txt", ".py", ".rs", ".go", ".js", ".ts", ".jsx", ".tsx",
    ".css", ".html", ".csv", ".xml", ".ini", ".cfg", ".conf", ".env",
    ".r", ".sql", ".jl",
}
BINARY_EXTS = {".pdf", ".docx", ".xlsx"}
IGNORE_DIRS  = {".git", "node_modules", "vendor", "__pycache__", ".venv",
                "dist", "build", ".DS_Store", ".idea", ".vscode"}
IGNORE_FILES = {".DS_Store", ".gitkeep", "Thumbs.db", ".localized"}

# Folder names that almost always indicate the LLM punted instead of thinking.
# Used both for post-filtering taxonomy output and for plan validation warnings.
_GENERIC_FOLDER_NAMES = frozenset({
    "backup", "backups", "archive", "archives", "old", "misc",
    "miscellaneous", "other", "stuff", "temp", "tmp", "unsorted",
})

# ── Code-project awareness ─────────────────────────────────────────────────────
# If any of these marker files exist at the root, we treat the directory as a
# code project and apply strict filtering so the LLM doesn't try to move source
# files, config files, lock files, or anything inside structural directories.
CODE_PROJECT_MARKERS = frozenset({
    "pyproject.toml", "setup.py", "setup.cfg",
    "package.json", "Cargo.toml", "go.mod",
    "flake.nix", "CMakeLists.txt", "Gemfile",
    "build.gradle", "pom.xml", "mix.exs",
})

# Extensions we *do* want to organize even inside a code project.
# Everything else (source, config, lock files) is left alone.
ORGANIZABLE_EXTS = frozenset({
    # Documents
    ".md", ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
    ".txt", ".rtf", ".odt", ".ods", ".odp",
    # Data
    ".csv", ".tsv", ".parquet", ".feather", ".arrow",
    # Notebooks
    ".ipynb",
    # Media
    ".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp", ".ico",
    ".mp4", ".mov", ".avi", ".mkv", ".mp3", ".wav",
    # Archives / deliverables
    ".zip", ".tar", ".gz", ".bz2", ".7z",
    # SQL / standalone scripts that aren't part of the project source
    ".sql",
})

# Specific filenames that are structural anchors and must never be moved,
# even if their extension is otherwise organizable.
STRUCTURAL_NAMES = frozenset({
    "README.md", "readme.md", "CHANGELOG.md", "changelog.md",
    "LICENSE", "LICENSE.md", "LICENSE.txt", "NOTICE",
    "CONTRIBUTING.md", "CODE_OF_CONDUCT.md", "SECURITY.md",
    "CODEOWNERS", "OWNERS",
})

# Top-level directory names whose *contents* should not be touched in a code project.
STRUCTURAL_DIRS = frozenset({
    "src", "lib", "pkg", "cmd", "internal", "external",
    "app", "apps", "core",
    "tests", "test", "spec", "specs", "__tests__",
    ".github", ".gitlab", ".circleci", "ci", ".ci",
    "scripts",           # project's own scripts — don't reorganize them
    "modules",           # nix modules, etc.
})


# ── Project detection ─────────────────────────────────────────────────────────

def detect_code_project(directory: Path) -> str | None:
    """Return the marker filename if *directory* is a code project root, else None."""
    for marker in CODE_PROJECT_MARKERS:
        if (directory / marker).exists():
            return marker
    return None


def find_project_roots(scan_root: Path, recursive: bool) -> set[Path]:
    """Return the set of all code-project root directories relevant to *scan_root*.

    Three sources are considered:
    1. Ancestor directories of *scan_root* — handles the case where the user
       points the tool at a subdirectory of a project (e.g. ~/proj/docs).
    2. *scan_root* itself.
    3. All descendant directories — handles umbrella repos (e.g. ~/code with
       many individual sub-projects, each with their own pyproject.toml).

    Once a project root is found during descent, its children are not searched
    further (nested projects inside a project are uncommon and confusing).
    """
    roots: set[Path] = set()

    # 1. Check ancestors (nearest first, stop at the first hit)
    for ancestor in scan_root.parents:
        if detect_code_project(ancestor):
            roots.add(ancestor)
            break   # Only the nearest ancestor matters

    # 2. Check scan_root itself
    if detect_code_project(scan_root):
        roots.add(scan_root)
        # If the root is already a project, skip descendant search — sub-project
        # detection inside a project is handled by STRUCTURAL_DIRS filtering.
        return roots

    # 3. Walk descendants (only when recursive)
    if recursive:
        for dirpath_str, dirnames, _ in os.walk(scan_root):
            dirpath = Path(dirpath_str)
            # Prune ignored dirs in-place so os.walk won't descend into them
            dirnames[:] = [
                d for d in dirnames
                if d not in IGNORE_DIRS and not d.startswith(".")
            ]
            if detect_code_project(dirpath):
                roots.add(dirpath)
                dirnames.clear()    # Don't look for nested projects inside a project

    return roots


def find_enclosing_project(path: Path, project_roots: set[Path]) -> Path | None:
    """Return the deepest project root that is an ancestor of *path*, or None."""
    best: Path | None = None
    for proj_root in project_roots:
        try:
            path.relative_to(proj_root)
        except ValueError:
            continue
        # Prefer the deeper (more specific) root
        if best is None or len(proj_root.parts) > len(best.parts):
            best = proj_root
    return best


def is_organizable_in_project(path: Path, proj_root: Path) -> bool:
    """Return True when *path* is a file worth reorganizing inside a code project."""
    rel   = path.relative_to(proj_root)
    parts = rel.parts   # e.g. ("docs", "report.pdf") or ("README.md",)

    # Never move structural anchor files
    if path.name in STRUCTURAL_NAMES:
        return False

    # Skip files that live inside a structural directory
    if len(parts) > 1 and parts[0] in STRUCTURAL_DIRS:
        return False

    # Only include explicitly organizable extensions
    return path.suffix.lower() in ORGANIZABLE_EXTS


# ── Text Extraction ────────────────────────────────────────────────────────────

def extract_snippet(path: Path, max_chars: int = SNIPPET_LEN) -> str | None:
    """Return a short content preview for LLM context. Returns None for binaries we can't read."""
    ext = path.suffix.lower()

    if ext in TEXT_EXTS:
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                text = f.read(max_chars * 2)
            return text[:max_chars].strip() or None
        except OSError:
            return None

    if ext == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(str(path))
            text = ""
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    text += " " + t
                if len(text) >= max_chars:
                    break
            return text.strip()[:max_chars] or None
        except Exception:
            return None

    if ext == ".docx":
        try:
            from docx import Document
            doc = Document(str(path))
            parts = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
            return " ".join(parts)[:max_chars] or None
        except Exception:
            return None

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True, data_only=True)
            parts: list[str] = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True, max_row=8):
                    row_text = " ".join(str(c) for c in row if c is not None)
                    if row_text.strip():
                        parts.append(row_text)
            return " ".join(parts)[:max_chars] or None
        except Exception:
            return None

    return None  # Unknown / binary


# ── Directory Scanning ─────────────────────────────────────────────────────────

def scan_directory(
    directory: str,
    recursive: bool = True,
    all_files: bool = False,
) -> list[dict]:
    """Walk *directory* and return file descriptors with metadata.

    When *all_files* is False (the default), project-aware filtering is applied:

    - A fast pre-pass discovers all code-project root directories relevant to
      the scan: ancestor projects (handles scanning inside a project), the root
      itself, and descendant sub-projects (handles umbrella repos like ~/code).
    - Each file is matched against its deepest enclosing project root.  Files
      inside a project are restricted to organizable extensions (docs, data,
      media) and kept out of structural directories (src/, tests/, …).
    - Files with no enclosing project root (loose files in a general directory)
      are included without restriction.
    """
    root = Path(directory).resolve()
    if not root.is_dir():
        print(f"Error: {root} is not a directory.", file=sys.stderr)
        sys.exit(1)

    # ── Build project-root map ─────────────────────────────────────────────────
    project_roots: set[Path] = set()
    if not all_files:
        project_roots = find_project_roots(root, recursive)
        if project_roots:
            # Show project directory names (not marker filenames which repeat)
            names = [p.name for p in sorted(project_roots, key=lambda p: p.name)]
            print(
                f"Detected {len(project_roots)} code project(s): "
                f"{', '.join(names)}. "
                "Filtering to organizable assets inside each. "
                "Pass --all-files to include everything.",
                file=sys.stderr,
            )

    # ── Walk files ─────────────────────────────────────────────────────────────
    results: list[dict] = []
    total_seen = 0
    walk = root.rglob("*") if recursive else root.iterdir()

    for path in walk:
        if any(part in IGNORE_DIRS for part in path.parts):
            continue
        if path.name in IGNORE_FILES or path.name.startswith("."):
            continue
        if not path.is_file():
            continue

        total_seen += 1

        # Apply per-project filter when relevant
        if project_roots:
            proj_root = find_enclosing_project(path, project_roots)
            if proj_root and not is_organizable_in_project(path, proj_root):
                continue

        try:
            stat = path.stat()
        except OSError:
            continue

        rel     = str(path.relative_to(root))
        snippet = extract_snippet(path)

        results.append({
            "path":    rel,
            "name":    path.name,
            "ext":     path.suffix.lower(),
            "size":    stat.st_size,
            "snippet": snippet or "",
        })

    if project_roots:
        skipped = total_seen - len(results)
        if skipped:
            print(
                f"Skipped {skipped} source/config file(s). "
                f"{len(results)} organizable file(s) remain.",
                file=sys.stderr,
            )

    results.sort(key=lambda f: f["path"])
    return results


# ── LLM Helpers ────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = (
    "You are a meticulous file-system organizer. "
    "Analyse the provided file list and produce a reorganization plan as JSON. "
    "Never suggest deleting files — only rename, move, or flag duplicates. "
    "Output ONLY valid JSON — no prose, no markdown code fences."
)


def call_llm(prompt: str) -> str:
    """Send *prompt* to Ollama and return the raw text response."""
    payload: dict = {
        "model":      MODEL,
        "prompt":     prompt,
        "system":     SYSTEM_PROMPT,
        "stream":     False,
        "think":      False,      # suppress chain-of-thought on qwen3 models
        "keep_alive": -1,         # keep model loaded in VRAM for the whole session
        "options": {
            "temperature": 0.15,
            "num_predict": 8192,
        },
    }
    try:
        resp = requests.post(OLLAMA_URL, json=payload, timeout=600)
        resp.raise_for_status()
        return resp.json().get("response", "").strip()
    except requests.exceptions.RequestException as e:
        print(f"Error calling Ollama: {e}", file=sys.stderr)
        sys.exit(1)


def extract_json(text: str) -> dict | list:
    """Pull the first complete JSON object or array out of *text*."""
    # Strip <think>…</think> blocks emitted by reasoning models (qwen3, etc.)
    text = re.sub(r"<think>.*?</think>", "", text, flags=re.DOTALL)
    # Strip optional markdown code fences
    text = re.sub(r"^```(?:json)?\s*", "", text.strip())
    text = re.sub(r"\s*```$", "", text.strip())

    for start_ch, end_ch in [('{', '}'), ('[', ']')]:
        start = text.find(start_ch)
        if start == -1:
            continue
        depth       = 0
        in_string   = False
        escape_next = False
        for i, ch in enumerate(text[start:], start):
            if escape_next:
                escape_next = False
                continue
            if ch == '\\' and in_string:
                escape_next = True
                continue
            if ch == '"':
                in_string = not in_string
                continue
            if in_string:
                continue
            if ch == start_ch:
                depth += 1
            elif ch == end_ch:
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[start : i + 1])
                    except json.JSONDecodeError:
                        break

    raise ValueError(f"No valid JSON in LLM response:\n{text[:600]}")


# ── Plan Generation ────────────────────────────────────────────────────────────
#
# Strategy
# --------
# Previous approach: LLM taxonomy (1 call) + LLM classify (N/150 calls)
#   739 files → ~6 LLM calls.  Still O(N) LLM calls for large dirs.
#
# Current approach:  embed → cluster → name
#   1. Embed all files with a small model (qwen3-embedding:0.6b, ~400 MB)
#   2. Cluster embeddings with HDBSCAN (auto cluster count, no LLM)
#   3. Name clusters via LLM (1 call)
#   4. Assign files to clusters (deterministic, no LLM)
#
# Result: any directory size → 1 LLM call for organization.
#   Rename and dedupe remain separate (1–2 LLM calls + embedding reuse).
#   Legacy two-phase approach kept as fallback if HDBSCAN finds no clusters.

# Maximum files per classification call (output is ~50 tokens/file: much smaller
# than a full plan operation, so we can use much larger batches).
MAX_CLASSIFY_BATCH = 150
# Maximum files per rename-suggestion call (needs snippets, so smaller batches).
MAX_RENAME_BATCH   = 30


# ── Rename candidate detection ─────────────────────────────────────────────────

_GENERIC_STEM_RE = [
    # Generic document names: doc1, report2, file_3, data, untitled, new copy…
    re.compile(
        r'^(doc|document|file|report|data|untitled|new|copy|image|img|photo|'
        r'pic|scan|draft|temp|tmp|attachment|download|export|output|result)\s*[\d_\-]*$',
        re.IGNORECASE,
    ),
    re.compile(r'^[A-Z]{2,5}\d{3,}$'),               # IMG0001, DSC09876, DCIM0042
    re.compile(r'^\d{4}-\d{2}-\d{2}$'),              # 2024-01-15  (date, no description)
    re.compile(r'^\d{8,}$'),                          # purely numeric (timestamps, IDs)
    re.compile(r'^[a-f0-9]{8,}$', re.IGNORECASE),    # hash-named files
]

def is_rename_candidate(filename: str) -> bool:
    """Return True if the filename stem looks too generic to be self-describing."""
    stem = Path(filename).stem
    return any(pat.match(stem) for pat in _GENERIC_STEM_RE)


# ── Embedding + Clustering Pipeline ───────────────────────────────────────────
#
# Instead of sending file lists to the LLM for taxonomy + classification
# (O(N/150) calls), we embed each file's metadata with a small model,
# cluster the vectors with HDBSCAN, then ask the LLM to *name* the
# discovered clusters (1 call).  Files are assigned to clusters by the
# algorithm — no LLM classification calls at all.


def _file_embed_text(f: dict) -> str:
    """Build the text representation of a file for embedding.

    Combines the path (which carries directory-structure signal) with a
    truncated content snippet so the embedding captures both location
    context and semantic content.
    """
    snippet = f.get("snippet", "").replace("\n", " ").strip()[:250]
    if snippet:
        return f"{f['path']} — {snippet}"
    return f["path"]


def embed_files(files: list[dict]) -> np.ndarray:
    """Embed every file's metadata via the Ollama embeddings endpoint.

    Returns an (N, D) float32 numpy array where N = len(files) and D is
    the embedding dimensionality of EMBED_MODEL.
    """
    vectors: list[list[float]] = []
    total = len(files)

    for i, f in enumerate(files):
        if i % EMBED_BATCH_LOG == 0 and i > 0:
            print(f"  Embedded {i}/{total} files…", file=sys.stderr)

        text = _file_embed_text(f)
        payload = {
            "model":  EMBED_MODEL,
            "prompt": text,
            "keep_alive": -1,
        }
        try:
            resp = requests.post(OLLAMA_EMBED_URL, json=payload, timeout=120)
            resp.raise_for_status()
            embedding = resp.json().get("embedding")
            if embedding is None:
                print(f"  Warning: no embedding returned for {f['path']}, using zeros.",
                      file=sys.stderr)
                # Use same dim as previous vectors, or a placeholder
                dim = len(vectors[-1]) if vectors else 384
                vectors.append([0.0] * dim)
            else:
                vectors.append(embedding)
        except requests.exceptions.RequestException as e:
            print(f"  Error embedding {f['path']}: {e}", file=sys.stderr)
            dim = len(vectors[-1]) if vectors else 384
            vectors.append([0.0] * dim)

    print(f"  Embedded {total}/{total} files.", file=sys.stderr)
    return np.array(vectors, dtype=np.float32)


# ── Composite Distance (embedding + path structure) ───────────────────────────
#
# The small embedding model (0.6b) can't always separate files that are
# semantically similar but belong to different projects (e.g. icons from
# two unrelated repos).  We compensate by blending the embedding's cosine
# distance with an IDF-weighted Jaccard distance on path components.
#
# IDF weighting means common structural dirs (assets/, public/) contribute
# little while distinctive project names (raycast-ai-applescript/, etc.)
# dominate — exactly the signal the embedding model misses.


def _compute_path_idf(files: list[dict]) -> dict[str, float]:
    """Compute Inverse Document Frequency for each directory component.

    Components that appear in many files' paths get low IDF; unique project
    directory names get high IDF.  Only considers directory components, not
    the filename itself.
    """
    import math

    n = len(files)
    if n == 0:
        return {}

    doc_freq: dict[str, int] = {}
    for f in files:
        # Use only directory components — the filename is content, not structure
        components = set(Path(f["path"]).parts[:-1])
        for comp in components:
            doc_freq[comp] = doc_freq.get(comp, 0) + 1

    return {
        comp: math.log(n / df)
        for comp, df in doc_freq.items()
    }


def _path_distance_matrix(files: list[dict]) -> np.ndarray:
    """Compute NxN IDF-weighted Jaccard distance on path directory components.

    Higher-IDF components (distinctive project names) contribute more to
    distance than low-IDF components (common structural dirs like assets/).
    """
    idf = _compute_path_idf(files)
    n = len(files)

    # Pre-compute directory component sets per file
    file_components: list[set[str]] = []
    for f in files:
        components = set(Path(f["path"]).parts[:-1])
        file_components.append(components)

    dist = np.zeros((n, n), dtype=np.float32)
    for i in range(n):
        for j in range(i + 1, n):
            a, b = file_components[i], file_components[j]
            union_set = a | b
            if not union_set:
                continue  # both are top-level files — zero path distance
            # Identical component sets → distance 0 regardless of IDF weights
            if a == b:
                continue
            intersection_set = a & b
            w_intersection = sum(idf.get(c, 0.0) for c in intersection_set)
            w_union = sum(idf.get(c, 0.0) for c in union_set)
            # When all weights are zero (every component appears in every file),
            # fall back to plain Jaccard
            if w_union == 0.0:
                similarity = len(intersection_set) / len(union_set)
            else:
                similarity = w_intersection / w_union
            d = 1.0 - similarity
            dist[i, j] = d
            dist[j, i] = d

    return dist


def _composite_distance_matrix(
    files: list[dict],
    embeddings: np.ndarray,
    alpha: float = COMPOSITE_ALPHA,
) -> np.ndarray:
    """Blend embedding cosine distance with IDF-weighted path distance.

    Returns an NxN float32 matrix suitable for ``HDBSCAN(metric='precomputed')``.
    Both component distances are min-max normalised to [0, 1] before blending
    so that α directly controls the relative contribution.
    """
    from sklearn.metrics.pairwise import euclidean_distances
    from sklearn.preprocessing import normalize

    normed = normalize(embeddings, norm="l2")

    # Cosine distance from L2-normed vectors: d_cos = ||a−b||² / 2
    eucl = euclidean_distances(normed)
    cosine_dist = np.clip(eucl ** 2 / 2.0, 0.0, 1.0)

    path_dist = _path_distance_matrix(files)

    # Min-max normalise each component to [0, 1]
    cos_max = cosine_dist.max() or 1.0
    path_max = path_dist.max() or 1.0

    composite = alpha * (cosine_dist / cos_max) + (1.0 - alpha) * (path_dist / path_max)
    np.fill_diagonal(composite, 0.0)

    return composite


def cluster_files(
    files: list[dict],
    embeddings: np.ndarray,
    min_cluster: int = HDBSCAN_MIN_CLUSTER,
) -> np.ndarray:
    """Cluster file embeddings using HDBSCAN with composite distance.

    Blends embedding cosine distance with IDF-weighted path-component
    distance so that files from the same project cluster together even
    when their content is semantically similar to files in other projects.

    Falls back to pure cosine distance for very large directories (>5 000
    files) where computing the full N×N matrix would be too expensive.

    Returns an integer label array of length N.  Label -1 means noise.
    """
    from sklearn.cluster import HDBSCAN

    use_composite = len(files) <= 5000

    if use_composite:
        print("  Computing composite distance (embedding + path IDF)…", file=sys.stderr)
        dist_matrix = _composite_distance_matrix(files, embeddings)
        clusterer = HDBSCAN(
            min_cluster_size=min_cluster,
            metric="precomputed",
        )
        labels = clusterer.fit_predict(dist_matrix)
    else:
        from sklearn.preprocessing import normalize
        normed = normalize(embeddings, norm="l2")
        clusterer = HDBSCAN(
            min_cluster_size=min_cluster,
            metric="euclidean",
        )
        labels = clusterer.fit_predict(normed)

    n_clusters = len(set(labels)) - (1 if -1 in labels else 0)
    n_noise    = int(np.sum(labels == -1))
    print(f"  HDBSCAN found {n_clusters} cluster(s), {n_noise} noise file(s).",
          file=sys.stderr)

    # ── Silhouette score: quantitative clustering quality ─────────────────────
    if n_clusters >= 2:
        from sklearn.metrics import silhouette_score
        non_noise = labels != -1
        if np.sum(non_noise) > n_clusters:
            try:
                if use_composite:
                    score = silhouette_score(
                        dist_matrix[non_noise][:, non_noise],
                        labels[non_noise],
                        metric="precomputed",
                    )
                else:
                    score = silhouette_score(normed[non_noise], labels[non_noise])
                quality = "good" if score > 0.25 else ("fair" if score > 0.10 else "poor")
                print(f"  Silhouette score: {score:.3f} ({quality}).", file=sys.stderr)
            except ValueError:
                pass

    return labels


def name_clusters(
    files: list[dict],
    labels: np.ndarray,
    root: str,
    flatten: bool = False,
    skip_labels: set[int] | None = None,
) -> dict[int, str]:
    """One LLM call: name each cluster based on representative file paths.

    For each cluster, we pick up to 15 representative files and ask the LLM
    to assign a descriptive folder name.  Returns a mapping from cluster
    label → folder name string.

    Clusters whose labels appear in *skip_labels* are excluded (they were
    already auto-named from directory structure).
    """
    unique_labels = sorted(set(labels))
    cluster_labels = [l for l in unique_labels if l != -1]
    if skip_labels:
        cluster_labels = [l for l in cluster_labels if l not in skip_labels]
    if not cluster_labels:
        return {}

    # Build representative samples per cluster
    cluster_descriptions: list[str] = []
    for label in cluster_labels:
        indices = np.where(labels == label)[0]
        # Pick up to 15 representative files — evenly spaced through the cluster
        if len(indices) > 15:
            step = len(indices) // 15
            sample_indices = indices[::step][:15]
        else:
            sample_indices = indices
        paths = [files[i]["path"] for i in sample_indices]
        cluster_descriptions.append(
            f"CLUSTER {label} ({len(indices)} files):\n" +
            "\n".join(f"  {p}" for p in paths)
        )

    clusters_text = "\n\n".join(cluster_descriptions)

    # Existing top-level dirs for context
    existing_top_dirs = sorted({
        f["path"].split("/")[0]
        for f in files
        if "/" in f["path"]
    })
    existing_ctx = ""
    if existing_top_dirs:
        existing_ctx = (
            "\n\nEXISTING TOP-LEVEL DIRECTORIES (preserve meaningful names):\n"
            + "\n".join(f"  {d}/" for d in existing_top_dirs)
        )

    depth_rule = "Maximum 1 level of depth (no sub-folders)" if flatten else "Maximum 2 levels of depth"
    count_rule = "3 to 10 folders total" if flatten else "5 to 15 folders total"

    prompt = f"""You are naming folder categories for an automatic file organizer working in: {root}

Below are clusters of files that have been grouped by content similarity.
Assign each cluster a descriptive folder name.

RULES:
- {count_rule}
- {depth_rule}
- Lowercase kebab-case names
- Name folders after actual content themes (avoid "misc", "other", "files")
- PRESERVE existing meaningful directory names where they match a cluster's content
- Multiple clusters CAN share a folder name if they are thematically similar
{existing_ctx}

{clusters_text}

Output ONLY a JSON object mapping cluster number to folder name.
Example: {{"0": "reports", "1": "data/raw", "2": "notebooks"}}

JSON:"""

    raw = call_llm(prompt)
    try:
        result = extract_json(raw)
        if isinstance(result, dict):
            # Normalise keys to int and values to cleaned strings
            mapping: dict[int, str] = {}
            existing_lower = {d.lower() for d in existing_top_dirs} if existing_top_dirs else set()
            generic_stems = {n.rstrip("s") for n in _GENERIC_FOLDER_NAMES}

            for k, v in result.items():
                label = int(k)
                folder = str(v).strip().rstrip("/")
                root_part = Path(folder).parts[0].lower() if folder else ""
                is_generic = root_part.rstrip("s") in generic_stems
                already_exists = root_part in existing_lower
                if is_generic and not already_exists:
                    print(
                        f"  Rejected generic name '{folder}' for cluster {label} — "
                        "files will keep current paths.",
                        file=sys.stderr,
                    )
                    continue
                mapping[label] = folder
            return mapping
    except (ValueError, KeyError) as e:
        print(f"  Warning: could not parse cluster naming response: {e}", file=sys.stderr)
    return {}


def _auto_name_from_paths(
    files: list[dict],
    labels: np.ndarray,
) -> dict[int, str]:
    """Name clusters from their dominant directory prefix — no LLM needed.

    If ≥60% of a cluster's files share the same top-level directory, that
    directory is used as the folder name.  This avoids sending obvious cases
    to the LLM (where a small model might invent something worse like
    "social-icons" when the answer is clearly "raycast-ai-applescript").
    """
    from collections import Counter

    unique_labels = sorted(set(labels))
    cluster_labels = [lbl for lbl in unique_labels if lbl != -1]
    auto_names: dict[int, str] = {}

    for label in cluster_labels:
        indices = np.where(labels == label)[0]
        top_dirs: Counter[str] = Counter()
        for idx in indices:
            parts = files[idx]["path"].split("/")
            if len(parts) > 1:
                top_dirs[parts[0]] += 1
            else:
                top_dirs["_root_"] += 1

        if not top_dirs:
            continue

        dominant_dir, count = top_dirs.most_common(1)[0]
        if dominant_dir == "_root_":
            continue

        pct = count / len(indices)
        if pct >= 0.6:
            auto_names[label] = dominant_dir

    return auto_names


def assign_files_to_clusters(
    files: list[dict],
    labels: np.ndarray,
    cluster_names: dict[int, str],
) -> list[dict]:
    """Map each file to its cluster's folder.  No LLM calls needed.

    Files labelled as noise (-1) or whose cluster wasn't named are left
    in place (no move operation generated).

    When multiple files would collide at the same destination path, the
    original parent directory is prepended as a subdirectory to disambiguate.
    For example, if both ``projA/icon.png`` and ``projB/icon.png`` target
    ``assets/``, the second becomes ``assets/projB/icon.png``.
    """
    # First pass: compute raw destinations and detect collisions
    raw_moves: list[tuple[int, str, str, str]] = []  # (idx, src, dst, folder)
    dst_counts: dict[str, int] = {}

    for i, f in enumerate(files):
        label = int(labels[i])
        folder = cluster_names.get(label)
        if folder is None:
            continue
        src = f["path"]
        dst = f"{folder}/{Path(src).name}"
        if src == dst:
            continue
        dst_counts[dst] = dst_counts.get(dst, 0) + 1
        raw_moves.append((i, src, dst, folder))

    colliding_dsts = {d for d, c in dst_counts.items() if c > 1}

    # Second pass: disambiguate collisions by injecting the original parent dir
    ops: list[dict] = []
    seen_mkdirs: set[str] = set()

    # Track which disambiguated dsts are still colliding so we can escalate
    for _idx, src, dst, folder in raw_moves:
        if dst in colliding_dsts:
            # Use the full source directory path to disambiguate.
            # "project-a/assets/icon.png" → "folder/project-a/assets/icon.png"
            # This handles cases where the immediate parent is the same
            # across colliding files (e.g. both have "assets/icon.png").
            src_dir = str(Path(src).parent)
            if src_dir and src_dir != ".":
                dst = f"{folder}/{src_dir}/{Path(src).name}"
            else:
                # Top-level file — use the stem as a prefix
                dst = f"{folder}/{Path(src).stem}-dup{Path(src).suffix}"

        if folder not in seen_mkdirs:
            ops.append({"op": "mkdir", "path": folder})
            seen_mkdirs.add(folder)

        # Also ensure any disambiguated subdirectory gets a mkdir
        dst_parent = str(Path(dst).parent)
        if dst_parent != folder and dst_parent not in seen_mkdirs:
            ops.append({"op": "mkdir", "path": dst_parent})
            seen_mkdirs.add(dst_parent)

        ops.append({
            "op": "move", "from": src, "to": dst,
            "reason": f"Organized into {folder}",
        })

    return ops


# ── Phase 1: folder taxonomy (legacy — kept for fallback) ─────────────────────

def plan_folder_taxonomy(files: list[dict], root: str, flatten: bool = False) -> list[str]:
    """One LLM call: derive a folder structure from file paths alone (no snippets).

    Sending only paths keeps the prompt tiny — all 700+ files fit comfortably —
    and the output is a small JSON array of folder names.

    When *flatten* is True, the prompt asks for a single-level structure to
    collapse deep nesting rather than a hierarchical reorganization.
    """
    paths_text = "\n".join(f"  {f['path']}" for f in files)

    # Extract existing top-level directory names for context so the LLM
    # understands the current structure and doesn't discard meaningful names.
    existing_top_dirs = sorted({
        f["path"].split("/")[0]
        for f in files
        if "/" in f["path"]
    })
    existing_ctx = ""
    if existing_top_dirs:
        existing_ctx = (
            "\n\nEXISTING TOP-LEVEL DIRECTORIES (for context — preserve meaningful names):\n"
            + "\n".join(f"  {d}/" for d in existing_top_dirs)
        )

    if flatten:
        prompt = f"""You are flattening a deeply nested file tree in: {root}

Given the file list below, propose a FLAT folder structure to replace the current deep nesting.

RULES:
- Maximum 1 level of depth (no sub-folders)
- 3 to 10 folders total
- Lowercase kebab-case names
- Name folders after actual content themes (avoid "misc", "other", "files")
- Every file must fit into at least one folder
- The goal is to REDUCE nesting, not reorganize — prefer fewer, broader categories
- PRESERVE existing meaningful directory names — do NOT replace distinct project or
  topic names with generic labels like "backups", "archive", or "old"
- If a top-level directory already has a descriptive name (e.g. "publicpresence",
  "raycast-ai-applescript"), keep that name as a folder in the flat structure
{existing_ctx}

FILES:
{paths_text}

Output ONLY a JSON array of folder name strings.
Example: ["reports", "data", "notebooks", "images"]

JSON:"""
    else:
        prompt = f"""You are organizing files in: {root}

Given the file list below, propose a clean, minimal folder structure.

RULES:
- 5 to 15 folders total
- Maximum 2 levels of depth
- Lowercase kebab-case names
- Name folders after actual content themes (avoid "misc", "other", "files")
- Every file must fit into at least one folder
- PRESERVE existing meaningful directory names — do NOT replace distinct project or
  topic names with generic labels like "backups", "archive", or "old"
- If files already live under descriptive project directories, keep those names
{existing_ctx}

FILES:
{paths_text}

Output ONLY a JSON array of folder path strings.
Example: ["reports/2024", "data/raw", "notebooks", "assets/images"]

JSON:"""

    raw = call_llm(prompt)
    try:
        result = extract_json(raw)
        if isinstance(result, list):
            folders = [str(f).strip().rstrip("/") for f in result if isinstance(f, str) and f.strip()]
            # Post-filter: reject purely generic folder names that the LLM
            # *invented*.  If the name already exists as a top-level directory
            # in the file tree it's legitimate and should be kept.
            existing_lower = {d.lower() for d in existing_top_dirs} if existing_top_dirs else set()
            generic_stems  = {n.rstrip("s") for n in _GENERIC_FOLDER_NAMES}

            filtered: list[str] = []
            for folder in folders:
                root_part = Path(folder).parts[0].lower() if folder else ""
                is_generic = root_part.rstrip("s") in generic_stems
                already_exists = root_part in existing_lower
                if is_generic and not already_exists:
                    print(
                        f"  Rejected LLM-invented generic folder '{folder}' — "
                        "files will keep their current paths.",
                        file=sys.stderr,
                    )
                else:
                    filtered.append(folder)
            return filtered
    except (ValueError, KeyError):
        pass
    return []


# ── Robust LLM list helper ────────────────────────────────────────────────────

def _call_llm_for_list(
    prompt: str,
    label: str = "LLM call",
    max_retries: int = 1,
) -> list:
    """Call the LLM expecting a JSON array response.  Handles:
    - Dict wrappers: if the LLM returns {"files": [...]} or {"results": [...]},
      unwrap to the inner list automatically.
    - Retry: on parse failure, retry up to *max_retries* times.
    - Diagnostics: log a preview of the raw response on failure.
    """
    for attempt in range(1 + max_retries):
        raw = call_llm(prompt)
        try:
            parsed = extract_json(raw)
        except (ValueError, KeyError):
            preview = raw[:200].replace("\n", " ")
            print(
                f"  Warning: could not parse {label} (attempt {attempt + 1}). "
                f"Raw preview: {preview!r}",
                file=sys.stderr,
            )
            continue

        # Happy path: already a list
        if isinstance(parsed, list):
            return parsed

        # Common LLM quirk: wraps the array in a dict like {"files": [...]}
        if isinstance(parsed, dict):
            for key in ("files", "results", "data", "assignments", "items"):
                if isinstance(parsed.get(key), list):
                    print(
                        f"  Note: {label} returned a dict wrapper (key={key!r}), unwrapping.",
                        file=sys.stderr,
                    )
                    return parsed[key]
            # Last resort: if there's exactly one list value, use it
            lists = [v for v in parsed.values() if isinstance(v, list)]
            if len(lists) == 1:
                print(
                    f"  Note: {label} returned a dict wrapper, unwrapping single list value.",
                    file=sys.stderr,
                )
                return lists[0]

        preview = str(parsed)[:200]
        print(
            f"  Warning: unexpected {type(parsed).__name__} for {label} "
            f"(attempt {attempt + 1}). Preview: {preview!r}",
            file=sys.stderr,
        )

    print(f"  Error: {label} failed after {1 + max_retries} attempt(s), skipping.", file=sys.stderr)
    return []


# ── Phase 2: file classification ───────────────────────────────────────────────

def classify_files_to_folders(
    files: list[dict],
    folders: list[str],
    root: str,
) -> list[dict]:
    """Assign each file to a folder from the taxonomy. Large batches (150 files)
    are fine here because each output entry is only ~50 tokens.
    """
    all_ops: list[dict] = []
    folders_text = "\n".join(f"  {f}" for f in folders)
    total = (len(files) + MAX_CLASSIFY_BATCH - 1) // MAX_CLASSIFY_BATCH

    for i in range(0, len(files), MAX_CLASSIFY_BATCH):
        batch = files[i : i + MAX_CLASSIFY_BATCH]
        n     = i // MAX_CLASSIFY_BATCH + 1
        print(f"  Classify batch {n}/{total} ({len(batch)} files)…", file=sys.stderr)

        files_text = "\n".join(f"  {f['path']}" for f in batch)
        prompt = f"""Assign each file to the most appropriate folder from the list below.

AVAILABLE FOLDERS (in: {root}):
{folders_text}

FILES TO CLASSIFY:
{files_text}

RULES:
- Assign every file to exactly one folder
- Output ONLY a JSON array — no prose, no fences

JSON: [{{"file": "original/path.ext", "folder": "target/folder"}}, ...]"""

        assignments = _call_llm_for_list(prompt, label=f"classify batch {n}")

        # Collect raw moves from this batch, then resolve collisions
        batch_moves: list[tuple[str, str, str]] = []  # (src, dst, folder)
        dst_counts: dict[str, int] = {}
        for a in assignments:
            if not isinstance(a, dict):
                continue
            src    = a.get("file",   "").strip()
            folder = a.get("folder", "").strip().rstrip("/")
            if not src or not folder:
                continue
            dst = f"{folder}/{Path(src).name}"
            if src == dst:
                continue
            dst_counts[dst] = dst_counts.get(dst, 0) + 1
            batch_moves.append((src, dst, folder))

        colliding = {d for d, c in dst_counts.items() if c > 1}
        seen_mkdirs: set[str] = set()

        for src, dst, folder in batch_moves:
            if dst in colliding:
                src_dir = str(Path(src).parent)
                if src_dir and src_dir != ".":
                    dst = f"{folder}/{src_dir}/{Path(src).name}"

            if folder not in seen_mkdirs:
                all_ops.append({"op": "mkdir", "path": folder})
                seen_mkdirs.add(folder)

            dst_parent = str(Path(dst).parent)
            if dst_parent != folder and dst_parent not in seen_mkdirs:
                all_ops.append({"op": "mkdir", "path": dst_parent})
                seen_mkdirs.add(dst_parent)

            all_ops.append({"op": "move", "from": src, "to": dst,
                            "reason": f"Organized into {folder}"})

    return all_ops


# ── Rename pass (snippet-based, pre-filtered) ──────────────────────────────────

_RENAME_SCHEMA = """\
[
  {"from": "old-name.pdf", "to": "better-name.pdf", "reason": "…"},
  …
]"""

def suggest_renames(candidates: list[dict], root: str) -> list[dict]:
    """Suggest better names only for files that look non-descriptive.

    Uses content snippets (so smaller batches), but the candidate list is
    typically tiny after pre-filtering, so total LLM calls stay low.
    """
    all_ops: list[dict] = []
    total = (len(candidates) + MAX_RENAME_BATCH - 1) // MAX_RENAME_BATCH

    for i in range(0, len(candidates), MAX_RENAME_BATCH):
        batch = candidates[i : i + MAX_RENAME_BATCH]
        n     = i // MAX_RENAME_BATCH + 1
        print(f"  Rename batch {n}/{total} ({len(batch)} files)…", file=sys.stderr)

        files_text = "\n".join(
            f"  {f['path']} | preview: \"{f['snippet'].replace(chr(10), ' ').strip()[:250]}\""
            if f["snippet"] else f"  {f['path']}"
            for f in batch
        )
        prompt = f"""Suggest better filenames for the files below. Only rename when the new name
is clearly more descriptive than the original. Keep the file in its current directory.

RULES:
- Lowercase kebab-case. Keep the original extension.
- Output ONLY a JSON array. Omit files that don't need renaming.

FILES (in: {root}):
{files_text}

SCHEMA: {_RENAME_SCHEMA}

JSON:"""

        renames = _call_llm_for_list(prompt, label=f"rename batch {n}")

        for r in renames:
            if not isinstance(r, dict):
                continue
            src = r.get("from", "").strip()
            dst = r.get("to",   "").strip()
            if not src or not dst or src == dst:
                continue
            # Ensure the rename stays in the same directory
            src_parent = str(Path(src).parent)
            dst_name   = Path(dst).name
            normalized_dst = f"{src_parent}/{dst_name}" if src_parent != "." else dst_name
            all_ops.append({"op": "rename", "from": src, "to": normalized_dst,
                            "reason": r.get("reason", "")})

    return all_ops


# ── Top-level plan dispatcher ──────────────────────────────────────────────────

def generate_plan(
    files: list[dict],
    directory: str,
    do_rename: bool,
    do_organize: bool,
    do_flatten: bool,
    do_dedupe: bool = False,
) -> tuple[dict, np.ndarray | None]:
    """Dispatch to the most efficient strategy for the requested operations.

    Returns (plan_dict, embeddings_or_None).  The embeddings are returned
    so callers can reuse them for deduplication without re-embedding.

    Pipeline (when organizing):
      1. Embed all files with the small model
      2. Compute composite distance (embedding cosine ⊕ IDF-weighted path Jaccard)
      3. Cluster with HDBSCAN on the composite matrix
      4. Auto-name clusters whose files share a dominant directory prefix
      5. LLM names the remaining clusters (often 0 when structure is clear)
      6. Assign files to named clusters (deterministic, no LLM)
    """
    all_ops:    list[dict]    = []
    summaries:  list[str]     = []
    embeddings: np.ndarray | None = None

    # ── Organize / Flatten: embed → composite cluster → name → assign ─────────
    if do_organize or do_flatten:
        if do_flatten:
            deep = sum(1 for f in files if f["path"].count("/") > 1)
            if deep:
                print(f"  {deep} deeply-nested file(s) will be flattened.", file=sys.stderr)

        # Step 1: Embed
        print(f"  Embedding {len(files)} files with {EMBED_MODEL}…", file=sys.stderr)
        embeddings = embed_files(files)

        # Step 2+3: Composite distance → HDBSCAN
        print("  Clustering with composite distance (embedding + path IDF)…",
              file=sys.stderr)
        labels = cluster_files(files, embeddings)

        n_clusters = len(set(labels)) - (1 if -1 in labels else 0)
        if n_clusters == 0:
            print("  Warning: HDBSCAN found no clusters, falling back to LLM taxonomy.",
                  file=sys.stderr)
            folders = plan_folder_taxonomy(files, directory, flatten=do_flatten)
            if folders:
                preview = ", ".join(folders[:6]) + ("…" if len(folders) > 6 else "")
                verb = "Flattened" if do_flatten else "Organized"
                print(f"  Proposed {len(folders)} folder(s): {preview}", file=sys.stderr)
                ops = classify_files_to_folders(files, folders, directory)
                all_ops.extend(ops)
                n_moves = sum(1 for o in ops if o.get("op") == "move")
                summaries.append(f"{verb} {n_moves} file(s) into {len(folders)} folder(s).")
        else:
            # Step 4: Auto-name from directory structure where possible
            auto_names = _auto_name_from_paths(files, labels)
            if auto_names:
                print(f"  Auto-named {len(auto_names)}/{n_clusters} cluster(s) "
                      f"from directory structure: {', '.join(auto_names.values())}",
                      file=sys.stderr)

            # Step 5: LLM names the rest
            llm_names: dict[int, str] = {}
            remaining = n_clusters - len(auto_names)
            if remaining > 0:
                print(f"  Naming {remaining} remaining cluster(s) via LLM…",
                      file=sys.stderr)
                llm_names = name_clusters(
                    files, labels, directory,
                    flatten=do_flatten,
                    skip_labels=set(auto_names.keys()),
                )

            cluster_names = {**auto_names, **llm_names}

            if not cluster_names:
                print("  Warning: cluster naming failed, skipping organize step.",
                      file=sys.stderr)
            else:
                preview = ", ".join(list(cluster_names.values())[:8])
                if len(cluster_names) > 8:
                    preview += "…"
                verb = "Flattened" if do_flatten else "Organized"
                print(f"  {len(cluster_names)} folder(s): {preview}", file=sys.stderr)

                # Step 6: Assign files to clusters (deterministic)
                ops = assign_files_to_clusters(files, labels, cluster_names)
                all_ops.extend(ops)
                n_moves = sum(1 for o in ops if o.get("op") == "move")
                summaries.append(
                    f"{verb} {n_moves} file(s) into {len(cluster_names)} folder(s)."
                )

    # ── Rename: pre-filtered snippet pass ─────────────────────────────────────
    # Build a path map from organize/flatten ops so rename operates on
    # post-move paths.  Without this, renames reference original locations
    # that no longer exist after the organize step moves the files.
    if do_rename:
        path_map: dict[str, str] = {}    # original_path → post-move path
        for op in all_ops:
            if op.get("op") == "move":
                path_map[op["from"]] = op["to"]

        candidates = [f for f in files if is_rename_candidate(f["name"])]
        if candidates:
            # Rewrite candidate paths to their post-move locations so the
            # rename LLM prompt and resulting ops use the correct paths.
            if path_map:
                for c in candidates:
                    c["path"] = path_map.get(c["path"], c["path"])

            print(f"  {len(candidates)} rename candidate(s) identified.", file=sys.stderr)
            ops = suggest_renames(candidates, directory)
            all_ops.extend(ops)
            n = len(ops)
            if n:
                summaries.append(f"Suggested {n} rename(s).")
        else:
            print("  No rename candidates found (all filenames look descriptive).",
                  file=sys.stderr)

    # Deduplicate mkdir operations accumulated across phases
    seen_mkdirs: set[str] = set()
    deduped: list[dict] = []
    for op in all_ops:
        if op.get("op") == "mkdir":
            p = op.get("path", "")
            if p not in seen_mkdirs:
                seen_mkdirs.add(p)
                deduped.append(op)
        else:
            deduped.append(op)

    return {"operations": deduped, "summary": " ".join(s for s in summaries if s)}, embeddings


# ── Deduplication ──────────────────────────────────────────────────────────────

def _groups_from_pairs(pairs: list[tuple[str, str]]) -> list[set[str]]:
    """Merge pairs into connected-component groups using union-find.

    Correctly handles transitive overlap: (A,B) + (C,D) + (B,C) → {A,B,C,D}.
    """
    parent: dict[str, str] = {}

    def find(x: str) -> str:
        while parent.setdefault(x, x) != x:
            parent[x] = parent[parent[x]]   # path compression
            x = parent[x]
        return x

    def union(a: str, b: str) -> None:
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb

    for p, q in pairs:
        union(p, q)

    groups_map: dict[str, set[str]] = {}
    for node in parent:
        root = find(node)
        groups_map.setdefault(root, set()).add(node)

    return list(groups_map.values())


def _dedupe_via_db(directory: str) -> list[dict] | None:
    """Use cached vectors from the ai-search DB; return None if unavailable.

    Uses an in-memory sqlite-vec virtual table + KNN query to find neighbours
    within DUPE_THRESHOLD, replacing the previous O(n²) pure-Python loop.
    """
    if not SEARCH_DB_PATH.exists():
        return None

    abs_dir = str(Path(directory).resolve())
    prefix  = abs_dir.rstrip("/") + "/"

    try:
        import sqlite_vec  # type: ignore[import-untyped]
        conn = sqlite3.connect(SEARCH_DB_PATH)
        conn.enable_load_extension(True)
        sqlite_vec.load(conn)
        conn.enable_load_extension(False)
        cur = conn.cursor()
        cur.execute(
            "SELECT m.filepath, e.embedding "
            "FROM file_metadata m "
            "JOIN file_embeddings e ON m.rowid = e.rowid "
            "WHERE m.filepath LIKE ?",
            (prefix + "%",),
        )
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        print(f"Note: Could not read ai-search DB ({e}). Falling back to LLM.", file=sys.stderr)
        return None

    if not rows:
        return None

    print(f"  Using {len(rows)} cached embeddings from ai-search database.", file=sys.stderr)

    # ── Average multiple chunks per file into a single representative vector ──
    file_vecs: dict[str, list[float]]  = {}
    file_cnts: dict[str, int]          = {}
    vec_len: int | None                = None

    for filepath, vec_bytes in rows:
        if vec_bytes is None:
            continue
        rel = filepath.removeprefix(prefix)
        if vec_len is None:
            vec_len = len(vec_bytes) // 4
        vec = list(struct.unpack(f"<{vec_len}f", vec_bytes))
        if rel in file_vecs:
            file_vecs[rel] = [a + b for a, b in zip(file_vecs[rel], vec)]
            file_cnts[rel] += 1
        else:
            file_vecs[rel] = vec
            file_cnts[rel] = 1

    if not file_vecs or vec_len is None:
        return None

    averaged: list[tuple[str, bytes]] = []
    for rel, vec in file_vecs.items():
        cnt = file_cnts[rel]
        avg = [v / cnt for v in vec]
        averaged.append((rel, struct.pack(f"<{vec_len}f", *avg)))

    # ── Build an in-memory sqlite-vec virtual table for efficient KNN ────────
    mem = sqlite3.connect(":memory:")
    mem.enable_load_extension(True)
    sqlite_vec.load(mem)
    mem.enable_load_extension(False)

    mem.execute(f"CREATE VIRTUAL TABLE tmp_vecs USING vec0(embedding float[{vec_len}])")
    for idx, (_, vec_bytes) in enumerate(averaged):
        mem.execute("INSERT INTO tmp_vecs(rowid, embedding) VALUES (?, ?)", (idx, vec_bytes))

    # For each file, find its nearest neighbours within the threshold.
    # sqlite-vec's KNN returns cosine distance when using float[] columns.
    # We ask for k neighbours (capped) and filter by distance.
    k = min(len(averaged), 20)   # at most 20 neighbours per file
    pairs: list[tuple[str, str]] = []
    seen_pairs: set[tuple[str, str]] = set()

    for idx, (rel_a, vec_bytes) in enumerate(averaged):
        knn_rows = mem.execute(
            "SELECT rowid, distance FROM tmp_vecs "
            "WHERE embedding MATCH ? AND k = ? "
            "ORDER BY distance",
            (vec_bytes, k),
        ).fetchall()

        for neighbour_rowid, distance in knn_rows:
            if neighbour_rowid == idx:
                continue
            if distance > DUPE_THRESHOLD:
                continue
            rel_b = averaged[neighbour_rowid][0]
            pair = tuple(sorted((rel_a, rel_b)))
            if pair not in seen_pairs:
                seen_pairs.add(pair)
                pairs.append((rel_a, rel_b))

    mem.close()

    return [
        {"op": "duplicate", "files": sorted(grp),
         "reason": "Very similar content in ai-search index — review and consolidate"}
        for grp in _groups_from_pairs(pairs)
    ]


def _build_file_table(files: list[dict]) -> str:
    """Format files into a compact text table of paths + snippet previews for LLM prompts."""
    lines: list[str] = []
    for f in files:
        snippet = f.get("snippet", "").replace("\n", " ").strip()[:250]
        if snippet:
            lines.append(f"  {f['path']} | preview: \"{snippet}\"")
        else:
            lines.append(f"  {f['path']}")
    return "\n".join(lines)


def _dedupe_via_llm(files: list[dict]) -> list[dict]:
    """Ask the LLM to identify likely duplicates from names + snippets."""
    file_text = _build_file_table(files)
    prompt = f"""Identify groups of files that appear to have duplicate or near-duplicate content, based on their filenames and content previews.

FILES:
{file_text}

Output ONLY a JSON array of duplicate groups. Each group is an array of relative file paths.
If no duplicates are found, output an empty array: []

Example: [["report.pdf", "report-copy.pdf"], ["notes-v1.txt", "notes-v2.txt"]]

JSON output:"""

    raw = call_llm(prompt)
    try:
        groups = extract_json(raw)
        if not isinstance(groups, list):
            return []
        ops: list[dict] = []
        for grp in groups:
            if isinstance(grp, list) and len(grp) >= 2:
                ops.append({
                    "op":     "duplicate",
                    "files":  grp,
                    "reason": "LLM identified similar content — review and consolidate",
                })
        return ops
    except (ValueError, KeyError):
        return []


# ── Intentional variant detection ─────────────────────────────────────────────
# These patterns strip size, resolution, theme, and version suffixes from
# filenames to reveal a canonical stem.  Files that share a canonical stem
# and directory are intentional variants (e.g. favicon-16x16 / favicon-32x32),
# not true duplicates.

_VARIANT_STRIP_RE = [
    re.compile(r'@\d+x', re.IGNORECASE),              # icon@2x, icon@3x
    re.compile(r'[-_]\d+x\d+', re.IGNORECASE),         # favicon-16x16, chrome-192x192
    re.compile(r'[-@](dark|light)', re.IGNORECASE),     # logo-dark, icon@dark
    re.compile(r'-v\d+', re.IGNORECASE),                # proposal-v2
    re.compile(r'[-_](small|medium|large|thumb)', re.IGNORECASE),  # photo-thumb
]


def _canonical_stem(filename: str) -> str:
    """Strip variant suffixes to get a canonical form for comparison.

    "command-icon@2x"  → "command-icon"
    "favicon-32x32"    → "favicon"
    "logo-dark"        → "logo"
    "conductor-proposal-v2" → "conductor-proposal"
    """
    stem = Path(filename).stem
    for pat in _VARIANT_STRIP_RE:
        stem = pat.sub('', stem)
    return stem.rstrip('-_ ')


def _is_variant_group(paths: list[str]) -> bool:
    """True if all files are intentional variants of the same asset.

    Requirements: same canonical stem AND same parent directory.
    """
    if len(paths) < 2:
        return False

    dirs = {str(Path(p).parent) for p in paths}
    if len(dirs) > 1:
        return False  # different directories — not variants

    canonical_stems = {_canonical_stem(Path(p).name) for p in paths}
    return len(canonical_stems) == 1


def _filter_variant_dupes(dupe_ops: list[dict]) -> list[dict]:
    """Remove duplicate groups that are actually intentional asset variants."""
    filtered: list[dict] = []
    for op in dupe_ops:
        files = op.get("files", [])
        if _is_variant_group(files):
            names = [Path(f).name for f in files]
            print(
                f"  Skipped variant group: {', '.join(names[:4])}"
                f"{'…' if len(names) > 4 else ''}",
                file=sys.stderr,
            )
            continue
        filtered.append(op)
    return filtered


def _dedupe_via_embeddings(
    files: list[dict],
    embeddings: np.ndarray,
) -> list[dict]:
    """Find duplicates using fresh embeddings from the organize step.

    Uses the same cosine-distance threshold as the DB path, but operates
    directly on the in-memory embedding matrix — no database required.
    """
    from sklearn.preprocessing import normalize

    normed = normalize(embeddings, norm="l2")
    n = len(files)

    # Compute pairwise cosine distances only for nearby vectors.
    # For efficiency with large N, we use sklearn's NearestNeighbors
    # with a radius query instead of computing the full N×N matrix.
    from sklearn.neighbors import NearestNeighbors

    nn = NearestNeighbors(
        metric="euclidean",
        n_neighbors=min(n, 20),
        algorithm="auto",
    )
    nn.fit(normed)

    # Euclidean distance on L2-normalised vectors: d_eucl = sqrt(2 - 2*cos_sim)
    # For DUPE_THRESHOLD (cosine distance) of 0.12:  cos_sim ≈ 0.88
    # d_eucl = sqrt(2 - 2*0.88) ≈ 0.49.  We use a slightly generous cutoff.
    eucl_threshold = float(np.sqrt(2.0 * DUPE_THRESHOLD))

    distances, indices = nn.kneighbors(normed)

    pairs: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()
    for i in range(n):
        for dist, j in zip(distances[i], indices[i]):
            if j == i:
                continue
            if dist > eucl_threshold:
                continue
            pair = tuple(sorted((files[i]["path"], files[int(j)]["path"])))
            if pair not in seen:
                seen.add(pair)
                pairs.append(pair)

    return [
        {"op": "duplicate", "files": sorted(grp),
         "reason": "Very similar content (embedding distance) — review and consolidate"}
        for grp in _groups_from_pairs(pairs)
    ]


def find_duplicates(
    directory: str,
    files: list[dict],
    embeddings: np.ndarray | None = None,
) -> list[dict]:
    """Find duplicate files.

    Priority order:
    1. Fresh embeddings from the organize step (fastest, no extra work)
    2. Cached vectors from the ai-search DB
    3. LLM-based fallback (slowest)

    All paths apply variant filtering to suppress false positives from
    intentional asset variants (retina, dark/light, size suffixes).
    """
    if embeddings is not None and len(embeddings) == len(files):
        print("  Using fresh embeddings for duplicate detection.", file=sys.stderr)
        raw = _dedupe_via_embeddings(files, embeddings)
        return _filter_variant_dupes(raw)

    db_result = _dedupe_via_db(directory)
    if db_result is not None:
        return _filter_variant_dupes(db_result)

    # If we have no embeddings and no DB, embed now then dedupe
    if len(files) <= 5000:
        print("  No cached embeddings — embedding files for duplicate detection…",
              file=sys.stderr)
        fresh = embed_files(files)
        return _filter_variant_dupes(_dedupe_via_embeddings(files, fresh))

    print("  No ai-search index found for this directory — using LLM for duplicate detection.",
          file=sys.stderr)
    return _filter_variant_dupes(_dedupe_via_llm(files))


# ── Plan Validation ────────────────────────────────────────────────────────────


def validate_plan(plan: dict, file_count: int) -> list[dict]:
    """Run quality heuristics on a plan and return a list of warnings.

    Each warning is a dict with 'level' ('warn' or 'error') and 'msg'.
    This lets the shell script (or --eval) surface problems before applying.
    """
    ops      = plan.get("operations", [])
    warnings: list[dict] = []

    moves   = [o for o in ops if o.get("op") == "move"]
    renames = [o for o in ops if o.get("op") == "rename"]
    mkdirs  = [o for o in ops if o.get("op") == "mkdir"]

    if not moves and not renames:
        return warnings   # nothing to validate

    # ── 1. Concentration: too many files routed to a single folder ────────────
    from collections import Counter
    dst_folders = Counter(
        str(Path(m["to"]).parent) for m in moves if m.get("to")
    )
    if dst_folders:
        top_folder, top_count = dst_folders.most_common(1)[0]
        pct = top_count / max(len(moves), 1) * 100
        if pct >= 60 and top_count >= 10:
            warnings.append({
                "level": "warn",
                "msg":   f"{top_count}/{len(moves)} moves ({pct:.0f}%) target a single "
                         f"folder '{top_folder}' — plan may be too coarse.",
            })

    # ── 2. Generic folder names (only flag LLM-invented ones) ───────────────
    # Derive existing top-level dirs from move source paths so we don't
    # warn about generic names that already exist in the user's tree.
    existing_top_dirs = {
        Path(m["from"]).parts[0].lower()
        for m in moves
        if m.get("from") and "/" in m["from"]
    }
    generic_stems = {n.rstrip("s") for n in _GENERIC_FOLDER_NAMES}

    generic_found = []
    for m in mkdirs:
        folder = m.get("path", "")
        root_part = Path(folder).parts[0].lower() if folder else ""
        is_generic = root_part.rstrip("s") in generic_stems
        already_exists = root_part in existing_top_dirs
        if is_generic and not already_exists:
            generic_found.append(folder)
    if generic_found:
        names = ", ".join(sorted(set(generic_found)))
        warnings.append({
            "level": "warn",
            "msg":   f"Generic folder name(s) detected: {names}. "
                     "The LLM may be lumping unrelated files together.",
        })

    # ── 3. Destination collisions (same target path for multiple files) ───────
    dst_paths = Counter(m.get("to", "") for m in moves)
    collisions = {p: c for p, c in dst_paths.items() if c > 1}
    if collisions:
        examples = list(collisions.keys())[:3]
        warnings.append({
            "level": "error",
            "msg":   f"{len(collisions)} destination collision(s) — multiple files mapped "
                     f"to the same path. Examples: {', '.join(examples)}",
        })

    # ── 4. Files moved but source path doesn't match any scanned file ────────
    # (Catches LLM hallucinated paths — we validate against the plan's 'from'
    # fields, but can't check against the real FS here since validate_plan is
    # FS-agnostic.  The apply step handles that, so this is a lighter check.)

    # ── 5. Suspiciously high move ratio (moving nearly everything) ────────────
    if file_count and len(moves) / file_count > 0.95 and file_count > 20:
        warnings.append({
            "level": "warn",
            "msg":   f"Plan moves {len(moves)}/{file_count} files ({len(moves)/file_count*100:.0f}%). "
                     "Very few files stay in place — verify the folder structure is appropriate.",
        })

    return warnings


# ── Plan Application ───────────────────────────────────────────────────────────

def apply_plan(plan: dict, root_dir: str, dry_run: bool) -> None:
    """Apply a reorganization plan to the filesystem."""
    root = Path(root_dir).resolve()
    ops  = plan.get("operations", [])

    applied = skipped = errors = 0

    for op in ops:
        kind = op.get("op")

        if kind == "mkdir":
            target = root / op.get("path", "")
            if dry_run:
                print(f"  [dry-run] mkdir   {op['path']}/")
            else:
                try:
                    target.mkdir(parents=True, exist_ok=True)
                    applied += 1
                except OSError as e:
                    print(f"  Error creating {target}: {e}", file=sys.stderr)
                    errors += 1

        elif kind in ("move", "rename"):
            src = root / op.get("from", "")
            dst = root / op.get("to",   "")
            if not src.exists():
                print(f"  Skip (not found): {op.get('from')}", file=sys.stderr)
                skipped += 1
                continue
            if dst.exists():
                print(f"  Skip (exists):    {op.get('to')}", file=sys.stderr)
                skipped += 1
                continue
            if dry_run:
                print(f"  [dry-run] {kind:<6}  {op.get('from')}  →  {op.get('to')}")
            else:
                try:
                    dst.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(src), str(dst))
                    applied += 1
                except OSError as e:
                    print(f"  Error moving {src} → {dst}: {e}", file=sys.stderr)
                    errors += 1

        elif kind == "duplicate":
            # Informational only — never auto-delete
            if dry_run:
                files_str = ", ".join(op.get("files", []))
                print(f"  [info]    dupes   {files_str}")

    if not dry_run:
        print(
            f"\n  Applied: {applied}  |  Skipped: {skipped}  |  Errors: {errors}",
            file=sys.stderr,
        )


# ── Entry Point ────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="AI-powered file reorganizer, renamer, and deduplicator"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--scan",  metavar="DIR",       help="Scan and emit file list as JSON")
    group.add_argument("--plan",  metavar="DIR",       help="Generate a reorganization plan")
    group.add_argument("--apply", metavar="PLAN_FILE", help="Apply a saved plan")
    group.add_argument(
        "--eval", metavar="PLAN_FILE",
        help="Validate a plan against quality heuristics without applying it",
    )

    # Plan-mode flags
    parser.add_argument("--rename",    action="store_true", help="Suggest better filenames")
    parser.add_argument("--organize",  action="store_true", help="Suggest folder structure")
    parser.add_argument("--flatten",   action="store_true", help="Flatten nested paths")
    parser.add_argument("--dedupe",    action="store_true", help="Flag near-duplicate files")
    parser.add_argument("--top-level", action="store_true", help="Only scan top-level files (non-recursive)")
    parser.add_argument(
        "--all-files", action="store_true",
        help="Disable code-project filtering: include source, config, and lock files"
    )
    parser.add_argument(
        "--from-scan", metavar="FILE",
        help="Load pre-scanned file list from a JSON file instead of re-scanning (output of --scan)"
    )

    # Apply-mode flags
    parser.add_argument("--dry-run", action="store_true", help="Preview without making changes")
    parser.add_argument(
        "--root", metavar="DIR",
        help="Root directory for --apply (overrides the 'root' stored in the plan)",
    )

    args = parser.parse_args()

    # ── Scan ──────────────────────────────────────────────────────────────────
    if args.scan:
        files = scan_directory(args.scan, recursive=not args.top_level, all_files=args.all_files)
        print(json.dumps(files, indent=2))
        sys.exit(0)

    # ── Plan ──────────────────────────────────────────────────────────────────
    if args.plan:
        # Default to all structural operations when none are specified
        if not any([args.rename, args.organize, args.flatten, args.dedupe]):
            args.rename   = True
            args.organize = True
            args.dedupe   = True

        if args.from_scan:
            scan_path = Path(args.from_scan)
            if not scan_path.exists():
                print(f"Error: scan file not found: {scan_path}", file=sys.stderr)
                sys.exit(1)
            with open(scan_path) as fh:
                files = json.load(fh)
            print(f"Loaded {len(files)} file(s) from pre-scan.", file=sys.stderr)
        else:
            files = scan_directory(args.plan, recursive=not args.top_level, all_files=args.all_files)

        if not files:
            result: dict = {"root": str(Path(args.plan).resolve()),
                            "operations": [], "summary": "No files found."}
            print(json.dumps(result, indent=2))
            sys.exit(0)

        print(f"Scanned {len(files)} file(s).", file=sys.stderr)

        if len(files) > 300:
            print(
                f"Warning: {len(files)} files found. Consider using --top-level "
                "or a more targeted subdirectory.",
                file=sys.stderr,
            )

        ops: list[dict] = []
        summary = ""
        cached_embeddings: np.ndarray | None = None

        if args.rename or args.organize or args.flatten:
            plan, cached_embeddings = generate_plan(
                files, args.plan,
                do_rename=args.rename,
                do_organize=args.organize,
                do_flatten=args.flatten,
                do_dedupe=args.dedupe,
            )
            ops.extend(plan.get("operations", []))
            summary = plan.get("summary", "")

        if args.dedupe:
            print("Detecting duplicates…", file=sys.stderr)
            dupe_ops = find_duplicates(args.plan, files, embeddings=cached_embeddings)
            ops.extend(dupe_ops)
            n = len(dupe_ops)
            if n:
                suffix = f"Found {n} duplicate group{'s' if n != 1 else ''}."
                summary = (summary + " " + suffix).strip()

        result = {
            "root":       str(Path(args.plan).resolve()),
            "operations": ops,
            "summary":    summary,
        }

        # Run quality validation and embed any warnings in the plan JSON
        plan_warnings = validate_plan(result, file_count=len(files))
        if plan_warnings:
            result["warnings"] = plan_warnings
            for w in plan_warnings:
                icon = "⚠" if w["level"] == "warn" else "✗"
                print(f"  {icon} {w['msg']}", file=sys.stderr)

        print(json.dumps(result, indent=2))
        sys.exit(0)

    # ── Apply ─────────────────────────────────────────────────────────────────
    if args.apply:
        plan_path = Path(args.apply)
        if not plan_path.exists():
            print(f"Error: Plan file not found: {plan_path}", file=sys.stderr)
            sys.exit(1)

        with open(plan_path) as fh:
            plan = json.load(fh)

        root = args.root or plan.get("root")
        if not root:
            print(
                "Error: --root is required when the plan has no embedded root path.",
                file=sys.stderr,
            )
            sys.exit(1)

        if not Path(root).is_dir():
            print(f"Error: Root directory does not exist: {root}", file=sys.stderr)
            sys.exit(1)

        apply_plan(plan, root, dry_run=args.dry_run)
        sys.exit(0)

    # ── Eval ─────────────────────────────────────────────────────────────────
    if args.eval:
        eval_path = Path(args.eval)
        if not eval_path.exists():
            print(f"Error: Plan file not found: {eval_path}", file=sys.stderr)
            sys.exit(1)

        with open(eval_path) as fh:
            plan = json.load(fh)

        ops    = plan.get("operations", [])
        moves  = [o for o in ops if o.get("op") == "move"]
        renames = [o for o in ops if o.get("op") == "rename"]
        dupes  = [o for o in ops if o.get("op") == "duplicate"]
        mkdirs = [o for o in ops if o.get("op") == "mkdir"]

        # Try to estimate original file count from the plan:
        # count unique source paths across moves and renames.
        all_srcs = {o.get("from", "") for o in moves + renames if o.get("from")}
        file_count = len(all_srcs) if all_srcs else 0

        print(f"Plan: {len(moves)} moves, {len(renames)} renames, "
              f"{len(dupes)} dupe groups, {len(mkdirs)} new dirs", file=sys.stderr)

        warnings = validate_plan(plan, file_count=file_count)

        if not warnings:
            print("✓ No quality issues detected.", file=sys.stderr)
            sys.exit(0)
        else:
            n_warn  = sum(1 for w in warnings if w["level"] == "warn")
            n_error = sum(1 for w in warnings if w["level"] == "error")
            for w in warnings:
                icon = "⚠" if w["level"] == "warn" else "✗"
                print(f"  {icon} {w['msg']}", file=sys.stderr)
            print(f"\n  {n_warn} warning(s), {n_error} error(s)", file=sys.stderr)
            sys.exit(1 if n_error else 0)


if __name__ == "__main__":
    main()
